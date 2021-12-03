'''
Python script to rename files
participantNames[]
fileNameList[]

'''

import openpyxl
from pathlib import Path
import os

participantNames = []
nameOfXlsxFile = "name.xlsx"
folderContainingCertfs = "certfs"


# Read names from xlsx file
xlsx_file = Path('.', nameOfXlsxFile)
wb_obj = openpyxl.load_workbook(xlsx_file) 
sheet = wb_obj.active
for row in sheet.iter_rows(2, sheet.max_row):
	participantNames.append((row[0].value))


fileNameList = os.listdir(folderContainingCertfs)
fileNameList.sort(key=lambda s: len(s))

idx = 0
dict = {}
for fileName in fileNameList:
	src = f"{folderContainingCertfs}/{fileName}"
	dst = f"{folderContainingCertfs}/{participantNames[idx]}.pdf"
	idx = idx+1
	print(src+ " = "+dst)
	os.rename(src, dst)
